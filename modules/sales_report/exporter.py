from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.exceptions import ExportError
from core.logger import get_logger
from modules.sales_report.analyzer import SalesAnalysisResult

logger = get_logger(__name__)

SUPPORTED_EXTENSION = ".xlsx"
MAX_COLUMN_WIDTH = 50
FREEZE_AND_FILTER_SHEETS = ("Detailed", "Staff", "Products", "Monthly")

DETAILED_NUMBER_FORMATS = {
    "date": "yyyy-mm-dd",
    "quantity": "0",
    "price": "#,##0",
    "sales": "#,##0",
}
GROUP_NUMBER_FORMATS = {
    "sales": "#,##0",
    "quantity": "0",
    "transaction_count": "0",
}
SUMMARY_NUMBER_FORMATS = {
    "total_sales": "#,##0",
    "total_quantity": "0",
    "transaction_count": "0",
    "average_sales_per_transaction": "#,##0.00",
}


class SalesReportExporter:
    def export(self, result: SalesAnalysisResult, output_path: str | Path) -> Path:
        path = Path(output_path)

        if path.suffix.lower() != SUPPORTED_EXTENSION:
            raise ExportError(
                f"Unsupported output format: {path.suffix}. Expected: {SUPPORTED_EXTENSION}"
            )

        try:
            path.parent.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            raise ExportError(f"Failed to create output directory: {path.parent}") from exc

        summary_display = result.summary.rename(columns={"metric": "Metric", "value": "Value"})
        sheets = [
            ("Summary", summary_display),
            ("Detailed", result.detailed),
            ("Staff", result.by_staff),
            ("Products", result.by_product),
            ("Monthly", result.by_month),
        ]

        try:
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                for sheet_name, frame in sheets:
                    frame.to_excel(writer, sheet_name=sheet_name, index=False)
                self._format_workbook(writer.book)
        except (OSError, ValueError) as exc:
            logger.exception("Failed to write Excel report: %s", path)
            raise ExportError(f"Failed to write Excel report: {path.name}") from exc

        return path

    def _format_workbook(self, workbook: Workbook) -> None:
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center")

        for worksheet in workbook.worksheets:
            self._style_header(worksheet, header_font, header_alignment)
            self._autofit_columns(worksheet)

            if worksheet.title in FREEZE_AND_FILTER_SHEETS:
                worksheet.freeze_panes = "A2"
                worksheet.auto_filter.ref = worksheet.dimensions

            if worksheet.title == "Detailed":
                self._apply_number_formats(worksheet, DETAILED_NUMBER_FORMATS)
            elif worksheet.title in ("Staff", "Products", "Monthly"):
                self._apply_number_formats(worksheet, GROUP_NUMBER_FORMATS)
            elif worksheet.title == "Summary":
                self._apply_summary_number_formats(worksheet)

    def _style_header(
        self, worksheet: Worksheet, font: Font, alignment: Alignment
    ) -> None:
        for cell in worksheet[1]:
            if cell.value is not None:
                cell.font = font
                cell.alignment = alignment

    def _autofit_columns(self, worksheet: Worksheet) -> None:
        for column_cells in worksheet.columns:
            lengths = [len(str(cell.value)) for cell in column_cells if cell.value is not None]
            if not lengths:
                continue
            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = min(
                max(lengths) + 2, MAX_COLUMN_WIDTH
            )

    def _column_index_map(self, worksheet: Worksheet) -> dict[str, int]:
        return {
            cell.value: index
            for index, cell in enumerate(worksheet[1], start=1)
            if cell.value is not None
        }

    def _apply_number_formats(
        self, worksheet: Worksheet, column_formats: dict[str, str]
    ) -> None:
        column_index = self._column_index_map(worksheet)
        for column_name, number_format in column_formats.items():
            col_idx = column_index.get(column_name)
            if col_idx is None:
                continue
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=col_idx).number_format = number_format

    def _apply_summary_number_formats(self, worksheet: Worksheet) -> None:
        for row in range(2, worksheet.max_row + 1):
            metric_cell = worksheet.cell(row=row, column=1)
            value_cell = worksheet.cell(row=row, column=2)
            number_format = SUMMARY_NUMBER_FORMATS.get(metric_cell.value)
            if number_format:
                value_cell.number_format = number_format
