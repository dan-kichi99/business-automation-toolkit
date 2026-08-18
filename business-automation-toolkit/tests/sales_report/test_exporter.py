import pandas as pd
import pytest
from openpyxl import load_workbook

from core.exceptions import ExportError
from modules.sales_report.analyzer import SalesAnalysisResult, SalesAnalyzer
from modules.sales_report.exporter import SalesReportExporter


def make_result() -> SalesAnalysisResult:
    data = pd.DataFrame(
        {
            "date": pd.to_datetime(["2026-08-01", "2026-08-01", "2026-08-02"]),
            "staff": ["Tanaka", "Sato", "Tanaka"],
            "product": ["Laptop", "Mouse", "Keyboard"],
            "quantity": [1, 2, 1],
            "price": [85000, 3500, 12000],
        }
    )
    return SalesAnalyzer().analyze(data)


@pytest.fixture
def exporter() -> SalesReportExporter:
    return SalesReportExporter()


@pytest.fixture
def result() -> SalesAnalysisResult:
    return make_result()


# ---- Normal cases ----


def test_export_creates_xlsx_file(exporter, result, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    exporter.export(result, output_path)

    assert output_path.exists()


def test_export_returns_output_path(exporter, result, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    returned_path = exporter.export(result, output_path)

    assert returned_path == output_path


def test_export_creates_expected_sheets_in_order(exporter, result, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    exporter.export(result, output_path)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Summary", "Detailed", "Staff", "Products", "Monthly"]


def test_export_summary_sheet_matches_analysis(exporter, result, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    exporter.export(result, output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook["Summary"]

    assert [cell.value for cell in worksheet[1]] == ["Metric", "Value"]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    assert rows[0] == ("total_sales", 104000)
    assert rows[1] == ("total_quantity", 4)
    assert rows[2] == ("transaction_count", 3)
    assert rows[3][0] == "average_sales_per_transaction"
    assert rows[3][1] == pytest.approx(104000 / 3)


def test_export_detailed_sheet_includes_sales(exporter, result, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    exporter.export(result, output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook["Detailed"]

    headers = [cell.value for cell in worksheet[1]]
    assert headers == ["date", "staff", "product", "quantity", "price", "sales"]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    assert rows[0][headers.index("sales")] == 85000


def test_export_staff_sheet_content(exporter, result, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    exporter.export(result, output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook["Staff"]

    headers = [cell.value for cell in worksheet[1]]
    assert headers == ["staff", "sales", "quantity", "transaction_count"]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    assert rows[0] == ("Tanaka", 97000, 2, 2)
    assert rows[1] == ("Sato", 7000, 2, 1)


def test_export_products_sheet_content(exporter, result, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    exporter.export(result, output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook["Products"]

    headers = [cell.value for cell in worksheet[1]]
    assert headers == ["product", "sales", "quantity", "transaction_count"]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    assert rows[0] == ("Laptop", 85000, 1, 1)


def test_export_monthly_sheet_content(exporter, result, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    exporter.export(result, output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook["Monthly"]

    headers = [cell.value for cell in worksheet[1]]
    assert headers == ["month", "sales", "quantity", "transaction_count"]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    assert rows[0] == ("2026-08", 104000, 4, 3)


def test_export_creates_missing_parent_directories(exporter, result, tmp_path):
    output_path = tmp_path / "nested" / "reports" / "sales_report.xlsx"

    exporter.export(result, output_path)

    assert output_path.exists()


def test_export_overwrites_existing_file(exporter, result, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    exporter.export(result, output_path)
    exporter.export(result, output_path)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Summary", "Detailed", "Staff", "Products", "Monthly"]


# ---- Formatting ----


def test_export_header_is_bold(exporter, result, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    exporter.export(result, output_path)

    workbook = load_workbook(output_path)
    for sheet_name in ["Summary", "Detailed", "Staff", "Products", "Monthly"]:
        assert workbook[sheet_name]["A1"].font.bold is True


def test_export_freeze_panes_on_data_sheets(exporter, result, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    exporter.export(result, output_path)

    workbook = load_workbook(output_path)
    for sheet_name in ["Detailed", "Staff", "Products", "Monthly"]:
        assert workbook[sheet_name].freeze_panes == "A2"


def test_export_date_number_format(exporter, result, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    exporter.export(result, output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook["Detailed"]
    assert worksheet.cell(row=2, column=1).number_format == "yyyy-mm-dd"


def test_export_sales_and_price_number_format(exporter, result, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    exporter.export(result, output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook["Detailed"]
    headers = [cell.value for cell in worksheet[1]]

    price_col = headers.index("price") + 1
    sales_col = headers.index("sales") + 1
    assert worksheet.cell(row=2, column=price_col).number_format == "#,##0"
    assert worksheet.cell(row=2, column=sales_col).number_format == "#,##0"


# ---- Abnormal cases ----


def test_export_unsupported_extension_raises_export_error(exporter, result, tmp_path):
    output_path = tmp_path / "sales_report.csv"

    with pytest.raises(ExportError):
        exporter.export(result, output_path)


def test_export_parent_path_is_file_raises_export_error(exporter, result, tmp_path):
    blocker = tmp_path / "blocker"
    blocker.write_text("not a directory", encoding="utf-8")
    output_path = blocker / "sales_report.xlsx"

    with pytest.raises(ExportError):
        exporter.export(result, output_path)


def test_export_output_path_is_directory_raises_export_error(exporter, result, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"
    output_path.mkdir()

    with pytest.raises(ExportError):
        exporter.export(result, output_path)


# ---- Immutability ----


def test_export_does_not_mutate_result(exporter, result, tmp_path):
    snapshots = {
        "detailed": result.detailed.copy(),
        "summary": result.summary.copy(),
        "by_staff": result.by_staff.copy(),
        "by_product": result.by_product.copy(),
        "by_month": result.by_month.copy(),
    }
    output_path = tmp_path / "sales_report.xlsx"

    exporter.export(result, output_path)

    pd.testing.assert_frame_equal(result.detailed, snapshots["detailed"])
    pd.testing.assert_frame_equal(result.summary, snapshots["summary"])
    pd.testing.assert_frame_equal(result.by_staff, snapshots["by_staff"])
    pd.testing.assert_frame_equal(result.by_product, snapshots["by_product"])
    pd.testing.assert_frame_equal(result.by_month, snapshots["by_month"])
