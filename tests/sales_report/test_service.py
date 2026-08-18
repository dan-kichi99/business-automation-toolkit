import io
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from core.exceptions import ExportError, FileReadError, ValidationError
from modules.sales_report.service import SalesReportService

CSV_CONTENT = (
    "date,staff,product,quantity,price\n"
    "2026-08-01,Tanaka,Laptop,1,85000\n"
    "2026-08-01,Sato,Mouse,2,3500\n"
    "2026-08-02,Tanaka,Keyboard,1,12000\n"
)


class RecordingComponents:
    def __init__(self):
        self.calls: list[str] = []

    def read(self, input_path):
        self.calls.append("read")
        return pd.read_csv(input_path)

    def validate(self, data):
        self.calls.append("validate")
        return data

    def analyze(self, data):
        self.calls.append("analyze")
        return data

    def export(self, result, output_path):
        self.calls.append("export")
        return Path(output_path)


@pytest.fixture
def service() -> SalesReportService:
    return SalesReportService()


@pytest.fixture
def csv_path(tmp_path) -> Path:
    path = tmp_path / "sales.csv"
    path.write_text(CSV_CONTENT, encoding="utf-8")
    return path


@pytest.fixture
def xlsx_path(tmp_path) -> Path:
    path = tmp_path / "sales.xlsx"
    df = pd.read_csv(io.StringIO(CSV_CONTENT))
    df.to_excel(path, index=False)
    return path


# ---- Normal cases ----


def test_generate_from_csv_creates_excel_report(service, csv_path, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    result = service.generate(csv_path, output_path)

    assert result == output_path
    assert output_path.exists()


def test_generate_from_xlsx_creates_excel_report(service, xlsx_path, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    result = service.generate(xlsx_path, output_path)

    assert result == output_path
    assert output_path.exists()


def test_generate_report_contains_expected_sheets(service, csv_path, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    service.generate(csv_path, output_path)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Summary", "Detailed", "Staff", "Products", "Monthly"]


def test_generate_report_summary_total_sales_matches(service, csv_path, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    service.generate(csv_path, output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook["Summary"]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    assert rows[0] == ("total_sales", 104000)


def test_generate_calls_components_in_order(tmp_path, csv_path):
    recorder = RecordingComponents()
    service = SalesReportService(
        reader=recorder, validator=recorder, analyzer=recorder, exporter=recorder
    )
    output_path = tmp_path / "sales_report.xlsx"

    result = service.generate(csv_path, output_path)

    assert recorder.calls == ["read", "validate", "analyze", "export"]
    assert result == output_path


# ---- Error propagation ----


def test_generate_missing_input_file_raises_file_read_error(service, tmp_path):
    output_path = tmp_path / "sales_report.xlsx"

    with pytest.raises(FileReadError):
        service.generate(tmp_path / "missing.csv", output_path)


def test_generate_unsupported_input_format_raises_file_read_error(service, tmp_path):
    input_path = tmp_path / "sales.txt"
    input_path.write_text("not a supported format", encoding="utf-8")
    output_path = tmp_path / "sales_report.xlsx"

    with pytest.raises(FileReadError):
        service.generate(input_path, output_path)


def test_generate_invalid_sales_data_raises_validation_error(service, tmp_path):
    input_path = tmp_path / "invalid.csv"
    input_path.write_text(
        "date,staff,product,quantity,price\n2026-08-01,Tanaka,Laptop,-1,85000\n",
        encoding="utf-8",
    )
    output_path = tmp_path / "sales_report.xlsx"

    with pytest.raises(ValidationError):
        service.generate(input_path, output_path)


def test_generate_unsupported_output_format_raises_export_error(service, csv_path, tmp_path):
    output_path = tmp_path / "sales_report.csv"

    with pytest.raises(ExportError):
        service.generate(csv_path, output_path)
